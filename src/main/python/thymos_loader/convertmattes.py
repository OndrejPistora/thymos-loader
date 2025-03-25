import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment
import polars as pl

def read_thymos_csv(file_path):
    metadata = {}
    data_rows = []
    header = None
    reading_state = "metadata"

    with open(file_path, 'r', newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            if reading_state == "metadata":
                if len(row) == 0:
                    reading_state = "header"
                elif len(row) >= 2:
                    metadata[row[0]] = row[1]
                elif len(row) == 1:
                    metadata[row[0]] = None
            elif reading_state == "header":
                header = row
                reading_state = "data"
            elif reading_state == "data":
                if len(row) == len(header):
                    data_rows.append([float(cell) if cell else 0.0 for cell in row])

    data = pl.DataFrame(data_rows, schema=[(col, pl.Float64) for col in header], orient="row")
    return metadata, data

def filter_only_loadcell2(data):
    # return only columns: time, position, loadcell2 renamed to force
    return data.select([
        pl.col("time"),
        pl.col("position"),
        pl.col("loadcell2").alias("force")
    ])

def convert_mattes(csv_files, output_excel_path):
    # sort by n from "*_n.csv"
    csv_files = sorted(csv_files, key=lambda x: int(os.path.basename(x).split("_")[-1].split(".")[0]))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Measurements"

    col_offset = 1

    for index, file_path in enumerate(csv_files, start=1):
        metadata, data = read_thymos_csv(file_path)
        data = filter_only_loadcell2(data)
        data = data.with_columns(pl.Series("deformation from F > 1N", [0.0] * len(data)))

        ws1.cell(row=1, column=col_offset, value=str(index)).alignment = Alignment(horizontal="center")
        ws1.merge_cells(start_row=1, start_column=col_offset, end_row=1, end_column=col_offset + len(data.columns) - 1)
        ws1.cell(row=2, column=col_offset, value="Measured values from testing machine").alignment = Alignment(horizontal="center")
        ws1.merge_cells(start_row=2, start_column=col_offset, end_row=2, end_column=col_offset + len(data.columns) - 1)

        # Header
        for i, col in enumerate(data.columns):
            ws1.cell(row=3, column=col_offset + i, value=col)
        
        # Units (customizable)
        units = {
            "time": "s",
            "position": "mm",
            "force": "N",
            "deformation from F > 1N": "mm",
        }
        for i, col in enumerate(data.columns):
            ws1.cell(row=4, column=col_offset + i, value=units.get(col, ""))

        # Data rows
        for r_idx, row in enumerate(data.iter_rows(named=True), start=5):
            for c_idx, col in enumerate(data.columns):
                ws1.cell(row=r_idx, column=col_offset + c_idx, value=row[col])

        col_offset += len(data.columns) + 1

    # Add second sheet
    ws2 = wb.create_sheet(title="objects")

    headers = [
        "Code", "Method", "W", "Number",
        "Dimensions, weight, before drying", "", "", "",
        "Dimensions, weight, after drying", "", "", "",
        "lo (mm)"
    ]
    headers2 = [
        "", "", "", "",
        "h_w", "l_w", "w_w", "m_w",
        "h_0", "l_0", "w_0", "m_0",
        ""
    ]
    headers3 = [
        "", "", "", "",
        "height (mm)", "length (mm)", "width (mm)", "mass (g)",
        "height (mm)", "length (mm)", "width (mm)", "mass (g)",
        ""
    ]
    ws2.append(headers)
    ws2.append(headers2)
    ws2.append(headers3)
    for i in range(1, 5):
        ws2.merge_cells(start_row=1, start_column=i, end_row=3, end_column=i)  # "Code", "Method", "W", "Number",
    ws2.merge_cells(start_row=1, start_column=13, end_row=3, end_column=13)  # Lo
    ws2.merge_cells(start_row=1, start_column=5, end_row=1, end_column=8)  # before drying
    ws2.merge_cells(start_row=1, start_column=9, end_row=1, end_column=12)  # after drying


    for i, filepath in enumerate(csv_files):
        filename = os.path.basename(filepath)
        row = [
            filename,
            3,  # Method
            0,  # W
            i + 1,  # Number
            "", "", "", "",  # before drying
            "", "", "", "",  # after drying
            ""  # length
        ]
        ws2.append(row)

    wb.save(output_excel_path)


if __name__ == "__main__":

    # TEST convert_mattes
    source_files = ["dev/test_2.csv", "dev/test_1.csv"]
    target_file = "dev/output.xlsx"
    convert_mattes(source_files, target_file)
    
    # # TEST read_thymos_csv
    # metadata, data = read_thymos_csv("dev/test_1.csv")
    # print(metadata)
    # print(data)
    